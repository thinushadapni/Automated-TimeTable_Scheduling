from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required

from django.db import models
import pandas as pd
from django.db.models import Q

from .models import Faculty, Course, Timetable, TimetableStatus, Student, Registration, Class
from .forms import (
    ClassForm,
    TimetableForm,
    FacultyUploadForm,
    StudentUploadForm,
    CourseUploadForm,
    RegistrationUploadForm,
    YearSemesterForm
)
from .validators import validate_timetable_constraints
from timetable_app.ga import run_ga_logic
from django.urls import reverse
# current_year="2025_even"
# current_semester="4"

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        selected_role = request.POST.get("role")  # Get the selected role
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Validate that the selected role is valid for the user
            valid_roles = ['TT_Coordinator', 'Department_Coordinator', 'faculty', 'student']
            if selected_role in valid_roles:
                login(request, user)
                request.session['selected_role'] = selected_role  # Store selected role in session
                return redirect("dashboard")
            else:
                messages.error(request, "Invalid role selected!")
                return redirect("login")
        else:
            messages.error(request, "Invalid credentials!")
            return redirect("login")

    return render(request, "login.html")

@login_required
def dashboard(request):
    return render(request, 'dashboard.html')

def upload_data(request, model, form_class, required_columns):
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            df = pd.read_csv(file) if file.name.endswith('.csv') else pd.read_excel(file)

            if not set(required_columns).issubset(df.columns):
                html_content = f"""
                <p>Invalid file format! Required columns: {', '.join(required_columns)}.</p>
                <a href='javascript:history.back()'>Go back to previous page</a>
                """
                return HttpResponse(html_content)

            existing_data = set(model.objects.values_list(*required_columns))

            new_records = []
            for _, row in df.iterrows():
                if tuple(row[col] for col in required_columns) not in existing_data:
                    record_data = {col: row[col] for col in required_columns}

                    # ✅ Special handling for Registration model
                    if model == Registration:
                        try:
                            record_data['stud_id'] = Student.objects.get(stud_id=row['stud_id'])
                        except Student.DoesNotExist:
                            html_content = f"""
                            <p>Student with ID {row['stud_id']} does not exist.</p>
                            <a href='javascript:history.back()'>Go back to previous page</a>
                            """
                            return HttpResponse(html_content)

                        try:
                            record_data['main_id'] = Class.objects.get(main_id=row['main_id'])
                        except Class.DoesNotExist:
                            html_content = f"""
                            <p>Class with ID {row['main_id']} does not exist.</p>
                            <a href='javascript:history.back()'>Go back to previous page</a>
                            """
                            return HttpResponse(html_content)
                            

                    new_records.append(model(**record_data))

            model.objects.bulk_create(new_records)

            html_content = f"""
            <p>{model.__name__} uploaded successfully!</p>
            <a href='javascript:history.back()'>Go back to previous page</a>
            """
            return HttpResponse(html_content)
    else:
        form = form_class()

    return render(request, f"upload_{model.__name__.lower()}.html", {'form': form})

# ✅ Updated Registration Upload View
def upload_registration(request):
    return upload_data(request, Registration, RegistrationUploadForm, ['stud_id', 'main_id'])

# Individual Upload Views
def upload_student(request):
    return upload_data(request, Student, StudentUploadForm, ['stud_id', 'name', 'department'])

def upload_faculty(request):
    return upload_data(request, Faculty, FacultyUploadForm, ['faculty_id', 'faculty_name', 'department'])

def upload_course(request):
    return upload_data(request, Course, CourseUploadForm, ['course_id', 'name', 'code', 'course_type', 'hours_per_week', 'offered_to'])


def add_class(request):
    """Add a new class."""
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Class added successfully!")
                return redirect('add_class')
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
    else:
        form = ClassForm()
    
    return render(request, 'add_class.html', {'form': form})

def select_year_semester(request):
    if request.method == 'POST':
        form = YearSemesterForm(request.POST)
        if form.is_valid():
            # Save year and sem to session
            request.session['current_year'] = form.cleaned_data['academic_year']
            request.session['current_semester'] = form.cleaned_data['semester']
            request.session['section'] = form.cleaned_data['section']
            request.session['dept'] = form.cleaned_data['dept']
            return redirect('add_timetable')
    else:
        form = YearSemesterForm()
    
    return render(request, 'select_year_semester.html', {'form': form})


def add_timetable(request):
    current_year = request.session.get('current_year')
    current_semester = request.session.get('current_semester')
    section = request.session.get('section')
    dept = request.session.get('dept')
    if not current_year or not current_semester:
        return redirect('select_year_semester')

    # Count Timetable entries for the given year and semester
    timetable_count = Timetable.objects.filter(
        main_id__academic_year=current_year,
        main_id__semester=current_semester
    ).filter(
        Q(main_id__section_id=section) | Q(main_id__section_id__isnull=True) | Q(main_id__section_id=""),
        Q(main_id__dept=dept) | Q(main_id__dept__isnull=True) | Q(main_id__dept="")
    ).count()

    # Get or create TimetableStatus based on count
    if timetable_count > 1:
        try:
            timetable_status = TimetableStatus.objects.get(
                academic_year=current_year,
                semester=current_semester,
                section=section,
                dept=dept
            )
        except TimetableStatus.DoesNotExist:
            # Fallback in case no status exists (edge case)
            timetable_status = TimetableStatus.objects.create(
                academic_year=current_year,
                semester=current_semester,
                section=section,
                dept=dept,
                status="tt_coordinator"
            )
    else:
        timetable_status, created = TimetableStatus.objects.get_or_create(
            academic_year=current_year,
            semester=current_semester,
            section=section,
            dept=dept,
            defaults={'status': 'tt_coordinator'}
        )

    print(request.user.role +','+timetable_status.status)
    
    # tt_courses = {'CLUB': 1, 'OE': 4, 'AE': 4}
    # dept_courses = {'ITT': 3, 'IAS': 2, 'DL LAB': 3, 'FS LAB': 3, 'SE LAB': 2, 'RP': 4, 'SS': 2, 'ASSO':1}

    tt_courses = {}
    dept_courses = {}

    #classes = Class.objects.filter(academic_year=current_year, semester=current_semester, section_id=section, dept=dept)
    classes = Class.objects.filter(
        academic_year=current_year,
        semester=current_semester
    ).filter(
        Q(section_id=section) | Q(section_id__isnull=True) | Q(section_id=""),
        Q(dept=dept) | Q(dept__isnull=True) | Q(dept="")
    )
    relevant_courses = set(cls.course for cls in classes)

    for course in relevant_courses:
        if course.course_type == 'tt':
            tt_courses[course.name] = course.hours_per_week
        elif course.course_type == 'dept':
            dept_courses[course.name] = course.hours_per_week
    print(tt_courses)
    
    # Get the list of already assigned courses with counts
    assigned_courses_count = dict(
        Timetable.objects.filter(
            main_id__academic_year=current_year,
            main_id__semester=current_semester
        ).filter(
            Q(main_id__section_id=section) | Q(main_id__section_id__isnull=True) | Q(main_id__section_id=""),
            Q(main_id__dept=dept) | Q(main_id__dept__isnull=True) | Q(main_id__dept="")
        ).values_list('main_id__course__name').annotate(count=models.Count('id'))
    )

    if 'ITT' in dept_courses:
        itt_slots = dept_courses.pop('ITT')
        dept_courses = {'ITT': itt_slots, **dept_courses}
    
    # Determine the current course to be assigned
    all_courses = list(tt_courses.keys()) + list(dept_courses.keys())
    
    current_course = None
    for course in all_courses:
        required_slots = tt_courses.get(course, dept_courses.get(course, 0))  # Get required slots
        assigned_slots = assigned_courses_count.get(course, 0)  # Get already assigned slots

        if assigned_slots < required_slots:
            current_course = course
            break  # Stop at the first unfulfilled course

    print(f"Current course to be assigned: {current_course}")

    if request.user.role == 'TT_Coordinator' and timetable_status.status != 'tt_coordinator':
        dashboard_url = reverse('dashboard')
        html_content = f"""Can't edit now. Please Wait. 
        <a href='{dashboard_url}'>Back to dashboard</a>"""
        return HttpResponse(html_content)

    if request.user.role == 'Department_Coordinator' and timetable_status.status not in ['dept_coordinator', 'ga_running']:
        dashboard_url = reverse('dashboard')
        html_content = f"""You can't edit now. Please wait. 
        <a href='{dashboard_url}'>Back to dashboard</a>"""
        return HttpResponse(html_content)

    if request.method == 'POST' and request.user.role in ['TT_Coordinator', 'Department_Coordinator']:
        form = TimetableForm(request.POST)

        if form.is_valid():
            main_id_obj = form.cleaned_data['main_id']  # This is a Class object
            selected_course = main_id_obj.course.name  # Get selected course name
            selected_days = form.cleaned_data['days']
            selected_slots = form.cleaned_data['slots']

            # Ensure the correct course is assigned in order
            if selected_course != current_course:
                html_content = f"""
                <p>Error: Please assign {current_course} before assigning {selected_course}.</p>
                <a href='javascript:history.back()'>Go back to previous page</a>
                """
                return HttpResponse(html_content)

            # Loop through selected days and slots
            for day in selected_days:
                for slot in selected_slots:
                    validation_error = validate_timetable_constraints(main_id_obj.main_id, day, slot, current_year, current_semester, section, dept)
                    if validation_error:
                        html_content = f"""
                        <p>validation_error : {validation_error}</p>
                        <a href='javascript:history.back()'>Go back to previous page</a>
                        """
                        return HttpResponse(html_content)

                    Timetable.objects.create(main_id=main_id_obj, day=day, slot=slot)

            # Recalculate assigned courses count after new entries
            assigned_courses_count = dict(
                Timetable.objects.filter(
                    main_id__academic_year=current_year,
                    main_id__semester=current_semester
                ).filter(
                    Q(main_id__section_id=section) | Q(main_id__section_id__isnull=True) | Q(main_id__section_id=""),
                    Q(main_id__dept=dept) | Q(main_id__dept__isnull=True) | Q(main_id__dept="")
                ).values_list('main_id__course__name').annotate(count=models.Count('id'))
            )
            
            # Check if all TT courses are assigned
            if timetable_status.status == 'tt_coordinator' and all(assigned_courses_count.get(course, 0) >= tt_courses[course] for course in tt_courses):
                timetable_status.status = 'dept_coordinator'

            # Check if all department courses are assigned
            elif timetable_status.status == 'dept_coordinator' and all(assigned_courses_count.get(course, 0) >= dept_courses[course] for course in dept_courses):
                timetable_status.status = 'ga_running'

            timetable_status.save()
            return redirect('add_timetable')

    else:
        form = TimetableForm()

    structured_timetable = defaultdict(lambda: defaultdict(list))  # Use a list instead of a single object
    days = set()
    slots = set()
    timetable = Timetable.objects.filter(
        main_id__academic_year=current_year,
        main_id__semester=current_semester
    ).filter(
        Q(main_id__section_id=section) | Q(main_id__section_id__isnull=True) | Q(main_id__section_id=""),
        Q(main_id__dept=dept) | Q(main_id__dept__isnull=True) | Q(main_id__dept="")
    )
    for entry in timetable:
        structured_timetable[entry.day][entry.slot].append(entry)  # Append to a list
        days.add(entry.day)
        slots.add(entry.slot)

    return render(request, 'add_timetable.html', {
        'form': form,
        'classes': classes,
        'current_course': current_course,
        'timetable_status': timetable_status,
        'current_year' : current_year,
        'current_semester' : current_semester,
        'section' : section,
        'dept' : dept,
        "timetable": structured_timetable,
        'days': sorted(days),   
        'slots': sorted(slots)
    })

@login_required
def run_genetic_algorithm(request):
    current_year = request.session.get('current_year')
    current_semester = request.session.get('current_semester')
    section = request.session.get('section')
    dept = request.session.get('dept')
    
    timetable_status = TimetableStatus.objects.filter(academic_year=current_year, semester=current_semester, section=section, dept=dept).first()

    if request.user.role != 'Department_Coordinator':
        html_content = f"""
        <p>You are not authorized to run the Genetic Algorithm.</p>
        <a href='javascript:history.back()'>Go back to previous page</a>
        """
        return HttpResponse(html_content)

    if timetable_status.status != 'ga_running':
        html_content = f"""
        <p>GA can't run yet!</p>
        <a href='javascript:history.back()'>Go back to previous page</a>
        """
        return HttpResponse(html_content)

    if not current_year or not current_semester:
        html_content = f"""
        <p>Missing year or semester.</p>
        <a href='javascript:history.back()'>Go back to previous page</a>
        """
        return HttpResponse(html_content)

    try:
        run_ga_logic(current_year, current_semester, section, dept)
        # timetable_status.status = 'completed'
        # timetable_status.save()
        return redirect('view_timetable')
    except Exception as e:
        html_content = f"""
        <p>Error running GA: {e}</p>
        <a href='javascript:history.back()'>Go back to previous page</a>
        """
        return HttpResponse(html_content)


from collections import defaultdict
def serialize_timetable(queryset):
    timetable = {}
    for entry in queryset:
        day = entry.day
        slot = entry.slot
        course_name = entry.main_id.course.name
        course_code = entry.main_id.course.code
        venue = entry.main_id.venue
        
        if day not in timetable:
            timetable[day] = {}
        timetable[day][slot] = timetable[day].get(slot, []) + [{'course_name': course_name, 'course_code': course_code, 'venue': venue}]
    return timetable

def view_timetable(request):
    classes = Class.objects.all()
    unique_years = classes.values_list('academic_year', flat=True).distinct()
    unique_semesters = classes.values_list('semester', flat=True).distinct()
    unique_section = classes.values_list('section_id', flat=True).distinct()
    unique_dept = classes.values_list('dept', flat=True).distinct()
    
    if request.method == "POST":
        user_input = request.POST.get("user_input", "").strip()  
        academic_year = request.POST.get('academic_year')
        semester = request.POST.get('semester')
        section = request.POST.get('section')
        dept = request.POST.get('dept')

        if not user_input:
            return render(request, "view_timetable.html", {"error": "Please enter a valid ID or 'admin'.","years": unique_years,"semesters": unique_semesters,"section" : unique_section,'dept' : unique_dept})

        # If Admin, show entire timetable
        if user_input.lower() == "admin":
            timetable = Timetable.objects.filter(
                main_id__academic_year=academic_year,
                main_id__semester=semester
            ).filter(
                Q(main_id__section_id=section) | Q(main_id__section_id__isnull=True) | Q(main_id__section_id=""),
                Q(main_id__dept=dept) | Q(main_id__dept__isnull=True) | Q(main_id__dept="")
            )
    
        # If Student, fetch registered courses
        elif Student.objects.filter(stud_id=user_input).exists():
            registered_courses = Registration.objects.filter(
                stud_id=user_input
            ).values_list('main_id', flat=True)

            timetable = Timetable.objects.filter(
                main_id__in=registered_courses,
                main_id__academic_year=academic_year,
                main_id__semester=semester
            )

        # If Faculty, fetch courses they handle
        elif Faculty.objects.filter(faculty_id=user_input).exists():
            faculty_courses = Class.objects.filter(
                faculty__faculty_id=user_input  # Use __ to filter ManyToManyField
            ).values_list('main_id', flat=True)
            timetable = Timetable.objects.filter(
                main_id__in=faculty_courses,
                main_id__academic_year=academic_year,
                main_id__semester=semester
            )

        # If Venue, fetch courses scheduled in that venue
        elif Class.objects.filter(venue=user_input).exists():
            class_main_ids = Class.objects.filter(
                venue=user_input
            ).values_list('main_id', flat=True)

            timetable = Timetable.objects.filter(
                main_id__in=class_main_ids,
                main_id__academic_year=academic_year,
                main_id__semester=semester
            ).filter(
                Q(main_id__dept=dept) | Q(main_id__dept__isnull=True) | Q(main_id__dept="")
            )

        else:
            return render(request, "view_timetable.html", {"error": "Invalid ID or venue entered.","years": unique_years,"semesters": unique_semesters,"section" : unique_section,'dept' : unique_dept})

        structured_timetable = defaultdict(lambda: defaultdict(lambda: None))
        days = set()
        slots = set()
        
        for entry in timetable:
            structured_timetable[entry.day][entry.slot] = structured_timetable[entry.day][entry.slot] or []  # Initialize as list if None
            structured_timetable[entry.day][entry.slot].append(entry)
            days.add(entry.day)
            slots.add(entry.slot)
            
        request.session['timetable_data'] = {
            "filtered_timetable": serialize_timetable(timetable),
            "days": list(days),
            "slots": list(slots)
        }
        
        return render(request, "view_timetable.html", {
            "timetable": structured_timetable,
            'classes': classes,
            "days": sorted(days),
            "slots": sorted(slots),
            "years": unique_years,
            "semesters": unique_semesters,
            "section": unique_section,  # Changed from 'section' to 'unique_section'
            "dept": unique_dept,       # Changed from 'dept' to 'unique_dept'
            "selected_section": section,  # Added to pre-select the chosen section
            "selected_dept": dept
        })

    return render(request, "view_timetable.html", {
        "years": unique_years,
        "semesters": unique_semesters,
        "section" : unique_section,
        'dept' : unique_dept
    })

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO

def download_timetable(request):
    session_data = request.session.get("timetable_data")
    if not session_data:
        # return HttpResponse("No filtered timetable data to export.", status=400)
        html_content = f"""
        <p>No filtered timetable data to export.</p>
        <a href='javascript:history.back()'>Go back to previous page</a>
        """
        return HttpResponse(html_content)

    timetable = session_data["filtered_timetable"]
    days = sorted(list(session_data["days"]))   
    slots = sorted(list(session_data["slots"])) 

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    ws.cell(row=1, column=1, value="Day / Slot").font = Font(bold=True)
    
    for idx, slot in enumerate(slots, start=2):
        ws.cell(row=1, column=idx, value=slot).font = Font(bold=True)

    for day_idx, day in enumerate(days, start=2):
        ws.cell(row=day_idx, column=1, value=day).font = Font(bold=True)
        for slot_idx, slot in enumerate(slots, start=2):
            entry = timetable.get(str(day), {}).get(str(slot))  #  keys in session are strings
            if entry:
                val = "\n".join(f"{e['course_name']} ({e['course_code']}) Venue: {e['venue']} /" for e in entry) if entry else "--"
            else:
                val = "--"
            cell = ws.cell(row=day_idx, column=slot_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for col in range(1, len(slots) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="timetable.xlsx"'
    return response

#python manage.py runserver